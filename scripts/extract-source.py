"""Extract matched occupation-total values from MHLW annual tables 6 and 7."""

from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

import openpyxl

YEARS = [2023, 2024, 2025]
OPENINGS_URL = "https://www.mhlw.go.jp/toukei/list/xls/114-1d-06.xlsx"
APPLICATIONS_URL = "https://www.mhlw.go.jp/toukei/list/xls/114-1d-07.xlsx"
SHEETS = {1: "a", 3: "f", 5: "t"}
EMPLOYMENTS = [
    {"id": "a", "name": "パートを含む常用", "short": "常用計"},
    {"id": "f", "name": "パートを除く常用", "short": "パート除く"},
    {"id": "t", "name": "常用的パートタイム", "short": "常用的パート"},
]
PREFECTURES = [
    ("JP-01", "北海道", "北海道"), ("JP-02", "青森", "東北"), ("JP-03", "岩手", "東北"),
    ("JP-04", "宮城", "東北"), ("JP-05", "秋田", "東北"), ("JP-06", "山形", "東北"),
    ("JP-07", "福島", "東北"), ("JP-08", "茨城", "関東"), ("JP-09", "栃木", "関東"),
    ("JP-10", "群馬", "関東"), ("JP-11", "埼玉", "関東"), ("JP-12", "千葉", "関東"),
    ("JP-13", "東京", "関東"), ("JP-14", "神奈川", "関東"), ("JP-15", "新潟", "北陸甲信越"),
    ("JP-16", "富山", "北陸甲信越"), ("JP-17", "石川", "北陸甲信越"), ("JP-18", "福井", "北陸甲信越"),
    ("JP-19", "山梨", "北陸甲信越"), ("JP-20", "長野", "北陸甲信越"), ("JP-21", "岐阜", "東海"),
    ("JP-22", "静岡", "東海"), ("JP-23", "愛知", "東海"), ("JP-24", "三重", "東海"),
    ("JP-25", "滋賀", "近畿"), ("JP-26", "京都", "近畿"), ("JP-27", "大阪", "近畿"),
    ("JP-28", "兵庫", "近畿"), ("JP-29", "奈良", "近畿"), ("JP-30", "和歌山", "近畿"),
    ("JP-31", "鳥取", "中国"), ("JP-32", "島根", "中国"), ("JP-33", "岡山", "中国"),
    ("JP-34", "広島", "中国"), ("JP-35", "山口", "中国"), ("JP-36", "徳島", "四国"),
    ("JP-37", "香川", "四国"), ("JP-38", "愛媛", "四国"), ("JP-39", "高知", "四国"),
    ("JP-40", "福岡", "九州・沖縄"), ("JP-41", "佐賀", "九州・沖縄"), ("JP-42", "長崎", "九州・沖縄"),
    ("JP-43", "熊本", "九州・沖縄"), ("JP-44", "大分", "九州・沖縄"), ("JP-45", "宮崎", "九州・沖縄"),
    ("JP-46", "鹿児島", "九州・沖縄"), ("JP-47", "沖縄", "九州・沖縄"),
]


def numeric(value: object) -> int | None:
    return int(value) if isinstance(value, (int, float)) and not isinstance(value, bool) else None


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load_openings(path: Path, place_ids: dict[str, str]) -> dict[tuple[str, str], list[int | None]]:
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    totals: dict[tuple[str, str], list[int | None]] = {}
    try:
        for sheet_index, employment in SHEETS.items():
            sheet = book.worksheets[sheet_index]
            place_id: str | None = None
            for row in sheet.iter_rows(min_row=3, max_col=5, values_only=True):
                if row[0] in place_ids:
                    place_id = place_ids[str(row[0])]
                if place_id and str(row[1] or "").strip() == "職業計":
                    totals[(place_id, employment)] = [numeric(value) for value in row[2:5]]
    finally:
        book.close()
    return totals


def load_applications(path: Path, place_ids: dict[str, str]) -> dict[tuple[str, str], list[int | None]]:
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    totals: dict[tuple[str, str], list[int | None]] = {}
    try:
        for sheet_index, employment in SHEETS.items():
            sheet = book.worksheets[sheet_index]
            place_id: str | None = None
            age: str | None = None
            for row in sheet.iter_rows(min_row=4, max_col=12, values_only=True):
                if row[0] in place_ids:
                    place_id = place_ids[str(row[0])]
                if row[1] is not None:
                    age = str(row[1]).strip()
                if place_id and age == "年齢計" and str(row[2] or "").strip() == "職業計":
                    totals[(place_id, employment)] = [numeric(row[offset]) for offset in (3, 6, 9)]
    finally:
        book.close()
    return totals


def main() -> None:
    if len(sys.argv) != 4:
        raise SystemExit("usage: extract-source.py OPENINGS.xlsx APPLICATIONS.xlsx OUTPUT_DIRECTORY")
    openings_path = Path(sys.argv[1])
    applications_path = Path(sys.argv[2])
    output_directory = Path(sys.argv[3])
    places = [{"id": "JP-00", "name": "全国", "region": "全国"}] + [
        {"id": item_id, "name": name, "region": region} for item_id, name, region in PREFECTURES
    ]
    place_ids = {"全国計": "JP-00"} | {
        f"{name}労働局": item_id for item_id, name, _region in PREFECTURES
    }
    openings = load_openings(openings_path, place_ids)
    applications = load_applications(applications_path, place_ids)
    expected_series = len(places) * len(EMPLOYMENTS)
    if len(openings) != expected_series or len(applications) != expected_series:
        raise ValueError(f"unexpected source series: {len(openings)=} {len(applications)=}")

    records: list[dict[str, object]] = []
    pair_count = 0
    employment_checks = {"openings": 0, "applications": 0}
    national_checks = {"openings": 0, "applications": 0}
    prefecture_ids = [place["id"] for place in places[1:]]
    for place in places:
        record: dict[str, object] = {"p": place["id"]}
        for employment in EMPLOYMENTS:
            employment_id = employment["id"]
            pairs = []
            for opening, application in zip(openings[(place["id"], employment_id)], applications[(place["id"], employment_id)], strict=True):
                if opening is None or application is None or opening < 0 or application <= 0:
                    raise ValueError(f"invalid value: {place['id']} {employment_id}")
                pairs.append([opening, application])
                pair_count += 1
            record[employment_id] = pairs
        records.append(record)

    for place in places:
        for year_index, year in enumerate(YEARS):
            for label, source in (("openings", openings), ("applications", applications)):
                all_value = source[(place["id"], "a")][year_index]
                full_value = source[(place["id"], "f")][year_index]
                part_value = source[(place["id"], "t")][year_index]
                if all_value != full_value + part_value:
                    raise ValueError(f"employment identity mismatch: {label} {place['id']} {year}")
                employment_checks[label] += 1

    for employment in EMPLOYMENTS:
        employment_id = employment["id"]
        for year_index, year in enumerate(YEARS):
            for label, source in (("openings", openings), ("applications", applications)):
                national = source[("JP-00", employment_id)][year_index]
                parts = [source[(place_id, employment_id)][year_index] for place_id in prefecture_ids]
                if national != sum(parts):
                    raise ValueError(f"national sum mismatch: {label} {employment_id} {year}")
                national_checks[label] += 1

    if pair_count != 432:
        raise ValueError(f"unexpected pair count: {pair_count}")
    index = {
        "schemaVersion": 1,
        "asOf": "2026-08-02",
        "edition": "2023〜2025年度（現行職業分類・職業計）",
        "years": YEARS,
        "placeCount": len(places),
        "prefectureCount": 47,
        "employmentCount": len(EMPLOYMENTS),
        "recordCount": len(records),
        "pairCount": pair_count,
        "sourceValueCount": pair_count * 2,
        "employmentIdentityChecked": employment_checks,
        "nationalSumChecked": national_checks,
        "places": places,
        "employments": EMPLOYMENTS,
        "sources": [
            {"kind": "openings", "url": OPENINGS_URL, "bytes": openings_path.stat().st_size, "sha256": sha256(openings_path)},
            {"kind": "applications", "url": APPLICATIONS_URL, "bytes": applications_path.stat().st_size, "sha256": sha256(applications_path)},
        ],
    }
    output_directory.mkdir(parents=True, exist_ok=True)
    (output_directory / "index.json").write_text(json.dumps(index, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    (output_directory / "ratios.json").write_text(json.dumps(records, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    print(json.dumps({"pairs": pair_count, "places": len(places), "source_values": pair_count * 2, "openings_sha256": index["sources"][0]["sha256"], "applications_sha256": index["sources"][1]["sha256"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
